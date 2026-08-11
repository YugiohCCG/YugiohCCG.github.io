#!/usr/bin/env python3
"""Build an illustrated Excel checklist for the current skipped CCG audit cases."""

from __future__ import annotations

import argparse
import json
import re
import tempfile
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageOps


ROOT = Path(__file__).resolve().parents[1]
CARDS_PATH = ROOT / "src" / "data" / "cards.json"
SKIPS_PATH = ROOT / "scripts" / "output" / "ccg_manual_input_skips.json"
MANUAL_LEDGER_PATH = ROOT / "scripts" / "output" / "omega_manual_test_ledger.json"
OUTPUT_DIR = ROOT / "scripts" / "output"


def output_path(card_count: int) -> Path:
    return OUTPUT_DIR / f"CCG_Manual_Input_Checklist_{card_count}.xlsx"

STATUS_OPTIONS = (
    "Not reviewed",
    "Needs ruling",
    "Needs live Omega test",
    "Engine limitation",
    "Change card text",
    "Change Lua",
    "Resolved",
    "Won't fix",
)

STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "be", "been", "by", "card", "cards",
    "current", "effect", "effects", "for", "from", "has", "have", "implementation",
    "in", "is", "it", "its", "needs", "not", "of", "on", "or", "the", "this",
    "to", "with", "without", "ruling", "test", "testing", "ui", "custom", "printed",
}

WORKSTREAM_TERMS = {
    "Card-pool and dependency definitions": {
        "add", "apply", "banish", "deck", "extra", "graveyard", "gy", "list", "lists",
        "monster", "name", "named", "reveal", "set", "spell", "summon", "trap",
    },
    "Copied, rewritten, and resolving effects": {
        "activation", "apply", "becomes", "chain", "copy", "effect", "gain", "gains",
        "negate", "resolve", "target",
    },
    "Co-link, counter, and zone topology": {
        "co-linked", "column", "counter", "linked", "link", "move", "points", "zone",
    },
    "Source wording and authoritative rulings": set(),
    "Damage and stat-layer behavior": {
        "atk", "battle", "damage", "def", "halve", "lp", "original", "reduce",
    },
    "Ownership and opponent-resource use": {
        "control", "deck", "material", "opponent", "owned", "tribute", "your",
    },
    "Timing, delayed state, and reset behavior": {
        "after", "battle", "during", "end", "if", "immediately", "next", "phase",
        "standby", "turn", "when",
    },
    "Summon, material, and selection UI": {
        "fusion", "link", "material", "normal", "ritual", "set", "special", "summon",
        "synchro", "tribute", "xyz",
    },
    "Other custom engine behavior": set(),
}

# Manual curation for cases where a terse historical audit label cannot identify
# the correct printed clause by lexical similarity alone. Each hint selects the
# complete sentence containing that exact fragment.
ISSUE_SENTENCE_HINTS: dict[int, tuple[str, ...]] = {
    254065048: ("loses ATK/DEF",),
    239245471: ("also treated as a \"Ataxia\" card", "If \"Azrynior, the Abundaence of Purity\" leaves"),
    224751741: ("also treated as a \"Taxis\" card", "If \"Dysmandr, the Depraevity of Worlds\" leaves"),
    222676270: ("When a card or effect is activated", "● Place", "● Set"),
    222782750: ("When a card or effect is activated", "● Add", "● Special Summon", "● Banish"),
    256831125: ("Link Rating", "gains the non-activated effects"),
    259841490: ("1+: You can move", "3+: During the Main Phase"),
    248801935: ("destroyed and sent from the field",),
    259670933: ("Must first be Special Summoned",),
    247831166: ("would Tribute a monster",),
    235051048: ("Cannot be Synchro Summoned", "Must be Special Summoned"),
    244921711: ("Cannot be Synchro Summoned", "Must be Special Summoned", "For the Special Summon", "Special Summon 1 WATER Monster", "shuffle up to 3"),
    249093610: ("Cannot be Synchro Summoned", "Must be Special Summoned"),
    259093260: ("For this card's Synchro Summon",),
    259219942: ("If this card is detached",),
    220749574: ("Must first be Special Summoned", "If this card is Special Summoned"),
    248946297: ("Add 1 \"Nautica\" monster",),
    218905439: ("Target 1 \"Nautica\" monster",),
    215105971: ("would be destroyed", "owner's control leaves"),
    213266433: ("can attack all face-down", "previously activated in this Chain"),
    248940511: ("used as Synchro Material",),
    259290896: ("cannot activate the effects of NATURE",),
    259377794: ("cannot be Summoned while you control a monster",),
}


def words(value: str) -> set[str]:
    return {
        word.casefold()
        for word in re.findall(r"[A-Za-z][A-Za-z'-]{2,}", value)
        if word.casefold() not in STOPWORDS
    }


def sentence_spans(text: str) -> list[tuple[int, int, str]]:
    spans: list[tuple[int, int, str]] = []
    start = 0
    for match in re.finditer(r"(?<=[.!?])(?:\s+|\n+)|\n+", text):
        end = match.start()
        if text[start:end].strip():
            left = start + len(text[start:end]) - len(text[start:end].lstrip())
            right = end - (len(text[start:end]) - len(text[start:end].rstrip()))
            spans.append((left, right, text[left:right]))
        start = match.end()
    if text[start:].strip():
        left = start + len(text[start:]) - len(text[start:].lstrip())
        right = len(text) - (len(text[start:]) - len(text[start:].rstrip()))
        spans.append((left, right, text[left:right]))
    return spans or [(0, len(text), text)]


def select_issue_spans(
    card_id: int,
    text: str,
    problem: str,
    implementation: str,
    workstream: str,
) -> list[tuple[int, int]]:
    sentences = sentence_spans(text)
    problem_lower = problem.casefold()

    curated = ISSUE_SENTENCE_HINTS.get(card_id, ())
    if curated:
        selected = [
            (start, end)
            for start, end, sentence in sentences
            if any(hint.casefold() in sentence.casefold() for hint in curated)
        ]
        if selected:
            return selected

    # High-confidence wording defects and unsupported branches.
    literal_hints: list[str] = []
    if "link rating" in problem_lower:
        literal_hints.append("link rating")
    if "set effect" in problem_lower or "set 1 spell/trap" in problem_lower:
        literal_hints.append("set 1")
    if "truncated" in problem_lower:
        return [(sentences[-1][0], sentences[-1][1])]
    if "flavor-only" in problem_lower or "metadata conflicts" in problem_lower:
        return [(0, len(text))]
    if "counter" in problem_lower:
        literal_hints.append("counter")
    if "co-link" in problem_lower or "linked" in problem_lower:
        literal_hints.extend(("co-link", "linked"))

    hinted = [
        (start, end)
        for start, end, sentence in sentences
        if any(hint in sentence.casefold() for hint in literal_hints)
    ]
    if hinted:
        return hinted

    issue_words = words(problem + " " + implementation)
    stream_words = WORKSTREAM_TERMS.get(workstream, set())
    scored: list[tuple[float, int, int, str]] = []
    for start, end, sentence in sentences:
        sentence_words = words(sentence)
        direct = len(sentence_words & issue_words)
        stream = len(sentence_words & stream_words)
        score = direct * 4 + stream
        lower = sentence.casefold()
        if any(term in problem_lower for term in ("copy", "copied", "rewritten")) and any(
            term in lower for term in ("apply", "becomes", "gain", "copy")
        ):
            score += 8
        if any(term in problem_lower for term in ("pool", "hardcoded", "named")) and any(
            term in lower for term in ("deck", "gy", "banish", "lists", "named", "reveal")
        ):
            score += 5
        if "summon" in problem_lower and "summon" in lower:
            score += 6
        if "damage" in problem_lower and "damage" in lower:
            score += 6
        scored.append((score, start, end, sentence))

    best = max((item[0] for item in scored), default=0)
    if best <= 0:
        return [(0, len(text))]
    # Keep tied/highly related clauses when one issue spans a bespoke procedure.
    threshold = max(1, best * 0.72)
    selected = [(start, end) for score, start, end, _ in scored if score >= threshold]
    return selected or [(scored[0][1], scored[0][2])]


def merge_spans(spans: list[tuple[int, int]]) -> list[tuple[int, int]]:
    merged: list[list[int]] = []
    for start, end in sorted(spans):
        if not merged or start > merged[-1][1]:
            merged.append([start, end])
        else:
            merged[-1][1] = max(merged[-1][1], end)
    return [(start, end) for start, end in merged]


def rich_effect_text(text: str, spans: list[tuple[int, int]]) -> CellRichText:
    normal = InlineFont(rFont="Aptos", sz=10, color="1F2937")
    issue = InlineFont(rFont="Aptos", sz=10, b=True, color="C00000")
    blocks: list[TextBlock] = []
    cursor = 0
    for start, end in spans:
        if cursor < start:
            blocks.append(TextBlock(normal, text[cursor:start]))
        blocks.append(TextBlock(issue, text[start:end]))
        cursor = end
    if cursor < len(text):
        blocks.append(TextBlock(normal, text[cursor:]))
    return CellRichText(*blocks)


def apply_excel_native_highlights(
    path: Path,
    sheet_title: str,
    spans_by_row: dict[int, list[tuple[int, int]]],
) -> None:
    """Use desktop Excel to write compatible per-character formatting.

    openpyxl can read rich text, but Excel 16 rejects the rich-text XML emitted
    by openpyxl 3.1.5. Generating plain cells first and formatting Characters
    through Excel produces a workbook that both applications open cleanly.
    """
    try:
        import win32com.client  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - this repository builds on Windows
        raise RuntimeError("Desktop Excel/pywin32 is required for native clause highlighting") from exc

    # DispatchEx creates an isolated automation instance and avoids touching a
    # workbook the user already has open. Dynamic dispatch also avoids pywin32
    # makepy-cache failures seen with EnsureDispatch on current Office builds.
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    book = None
    try:
        book = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
        sheet = book.Worksheets(sheet_title)
        for row, spans in spans_by_row.items():
            cell = sheet.Cells(row, 7)
            for start, end in spans:
                font = cell.GetCharacters(start + 1, end - start).Font
                font.Bold = True
                font.Color = 192  # VBA RGB(192, 0, 0)
        book.Save()
    finally:
        if book is not None:
            book.Close(SaveChanges=False)
        excel.Quit()


def thumbnail(source: Path, destination: Path) -> None:
    with Image.open(source) as image:
        image = ImageOps.exif_transpose(image).convert("RGB")
        image.thumbnail((260, 380), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", image.size, "white")
        canvas.paste(image)
        canvas.save(destination, "JPEG", quality=70, optimize=True, progressive=True)


def resolution_action(test: dict[str, Any], static_verdict: str) -> str:
    checklist = test.get("checklist", [])
    workstream = str(test.get("workstream", ""))
    category_action = {
        "Card-pool and dependency definitions": "Define the exact legal pool (name, set code, listed text, or explicit IDs), including how future cards enter it.",
        "Copied, rewritten, and resolving effects": "Define copied-effect ownership, costs, targets, count limits, and interruption behavior, then verify each branch in Omega.",
        "Co-link, counter, and zone topology": "Build the positive, illegal-zone, moved-card, and counter-boundary board states in Omega and record the UI/result.",
        "Source wording and authoritative rulings": "Replace the ambiguous or malformed wording with definitive PSCT-like text before changing Lua.",
        "Damage and stat-layer behavior": "Adopt the intended damage/stat-layer ruling and verify calculation, reset, and boundary values in Omega.",
        "Ownership and opponent-resource use": "Define owner/controller/location legality and confirm the opponent-resource selection UI and movement reasons in Omega.",
        "Timing, delayed state, and reset behavior": "Define the exact trigger event, duration, reset, and missed-timing behavior, then test interruption cases.",
        "Summon, material, and selection UI": "Define the exact summon/material procedure and validate legal, illegal, full-zone, cancel, and revival cases in Omega.",
        "Other custom engine behavior": "Reproduce the custom mechanic in Omega, record the closest legal and illegal cases, and revise text or Lua if behavior differs.",
    }.get(workstream, "Adopt the intended ruling and verify it in Omega.")
    prefix = (
        "Decide whether to revise the printed effect or implement an engine-supported alternative. "
        if static_verdict == "UNSUPPORTED"
        else "Adopt the intended ruling, then compare the Lua literally against that ruling. "
    )
    steps = "\n".join(f"• {step}" for step in checklist[:4])
    return prefix + category_action + ("\n" + steps if steps else "")


def explain_issue(problem: str, workstream: str, static_verdict: str) -> str:
    if static_verdict == "UNSUPPORTED":
        why = "The bundled Omega API does not expose a faithful implementation path for at least one printed branch. A card-text change or an engine-supported redesign is required."
    else:
        why = {
            "Card-pool and dependency definitions": "The implementation depends on an interpreted or explicit card pool that static comparison cannot prove is the designer's intended set, especially for future cards.",
            "Copied, rewritten, and resolving effects": "Omega's copy/rewrite helpers do not automatically prove original costs, targets, ownership, count limits, and interruption semantics.",
            "Co-link, counter, and zone topology": "The Lua structure is plausible, but static checks cannot prove live zone masks, movement prompts, co-link transitions, or counter traversal.",
            "Source wording and authoritative rulings": "The printed wording is ambiguous, malformed, truncated, or inconsistent, so more than one script behavior could reasonably follow from it.",
            "Damage and stat-layer behavior": "The intended calculation layer, rounding, duration, or boundary behavior is not uniquely established by static script comparison.",
            "Ownership and opponent-resource use": "The effect uses opponent-owned or unusually located resources; ownership, controller, movement reason, and selection legality need an authoritative ruling/live check.",
            "Timing, delayed state, and reset behavior": "The exact event snapshot, trigger timing, delayed state, or reset cannot be certified without a live scenario.",
            "Summon, material, and selection UI": "The effect uses a bespoke summon/material/selection procedure whose actual prompts and legality boundaries require an interactive Omega duel.",
            "Other custom engine behavior": "The behavior depends on a bespoke engine interaction that syntax and callback harnesses cannot certify.",
        }.get(workstream, "The remaining behavior cannot be proven by static analysis alone.")
    return f"Audit flag: {problem}\n\nWhy unresolved: {why}"


def build_workbook() -> tuple[Workbook, dict[str, int]]:
    cards = json.loads(CARDS_PATH.read_text(encoding="utf-8-sig"))
    cards_by_id = {int(card["passcode"]): card for card in cards}
    skips = json.loads(SKIPS_PATH.read_text(encoding="utf-8"))
    manual = json.loads(MANUAL_LEDGER_PATH.read_text(encoding="utf-8"))
    test_by_id = {int(test["card_id"]): test for test in manual["tests"]}
    rows = skips["unresolved"]
    workbook_path = output_path(len(rows))
    checklist_title = f"{len(rows)} Card Checklist"

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    ws = wb.create_sheet(checklist_title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:S{len(rows) + 1}"

    headers = [
        "Status", "Done?", "Card", "Card name", "Passcode", "Archetype", "Full printed effect text",
        "Issue effect text (highlighted above)", "What is wrong / unknown", "Current audited implementation",
        "What needs to be done", "Evidence required to resolve", "Workstream", "Static verdict", "Script",
        "Website image source", "Text SHA-256", "Script SHA-256", "Reviewer notes / adopted ruling",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(name="Aptos Display", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 38

    widths = [22, 10, 19, 31, 13, 22, 68, 55, 48, 52, 58, 52, 31, 18, 19, 48, 18, 18, 52]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    issue_fill = PatternFill("solid", fgColor="FFF2CC")
    unresolved_fill = PatternFill("solid", fgColor="FCE4D6")
    engine_fill = PatternFill("solid", fgColor="E4DFEC")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    highlight_spans_by_row: dict[int, list[tuple[int, int]]] = {}

    with tempfile.TemporaryDirectory(prefix="ccg-checklist-") as temp_name:
        temp_dir = Path(temp_name)
        for excel_row, item in enumerate(rows, 2):
            card_id = int(item["card_id"])
            card = cards_by_id[card_id]
            test = test_by_id[card_id]
            text = str(card.get("text", ""))
            implementation = str(test.get("audited_implementation", ""))
            spans = merge_spans(select_issue_spans(card_id, text, item["problem"], implementation, item["workstream"]))
            issue_text = "\n\n".join(text[start:end] for start, end in spans)
            highlight_spans_by_row[excel_row] = spans
            image_relative = str(card.get("image", ""))
            image_path = ROOT / "public" / image_relative.lstrip("/")
            if not image_path.exists():
                raise FileNotFoundError(f"{card_id}: website image missing: {image_path}")

            if item["static_verdict"] == "UNSUPPORTED":
                initial_status = "Engine limitation"
            elif item["workstream"] in {
                "Card-pool and dependency definitions",
                "Source wording and authoritative rulings",
            }:
                initial_status = "Needs ruling"
            else:
                initial_status = "Needs live Omega test"
            values = [
                initial_status,
                "☐",
                "",
                card["name"],
                card_id,
                card.get("archetype") or "Standalone",
                "",
                issue_text,
                explain_issue(item["problem"], item["workstream"], item["static_verdict"]),
                implementation,
                resolution_action(test, item["static_verdict"]),
                test.get("acceptance_condition", ""),
                item["workstream"],
                item["static_verdict"],
                test.get("script", f"c{card_id}.lua"),
                image_relative,
                item.get("text_sha256", ""),
                item.get("script_sha256", ""),
                "",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(excel_row, col, value)
                cell.border = border
                cell.fill = normal_fill
                cell.font = Font(name="Aptos", size=10, color="1F2937")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Excel-native rich formatting is applied after the plain workbook
            # has been saved; see apply_excel_native_highlights().
            ws.cell(excel_row, 7).value = text
            ws.cell(excel_row, 8).fill = issue_fill
            ws.cell(excel_row, 8).font = Font(name="Aptos", size=10, bold=True, color="C00000")
            ws.cell(excel_row, 9).fill = unresolved_fill
            ws.cell(excel_row, 14).fill = engine_fill if item["static_verdict"] == "UNSUPPORTED" else unresolved_fill
            ws.cell(excel_row, 5).number_format = "0"
            ws.cell(excel_row, 16).hyperlink = image_path.resolve().as_uri()
            ws.cell(excel_row, 16).style = "Hyperlink"
            ws.cell(excel_row, 2).font = Font(name="Segoe UI Symbol", size=16, color="7F6000")
            ws.cell(excel_row, 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[excel_row].height = 205

            thumb_path = temp_dir / f"{card_id}.jpg"
            thumbnail(image_path, thumb_path)
            xl_image = XLImage(str(thumb_path))
            xl_image.width = 128
            xl_image.height = 187
            ws.add_image(xl_image, f"C{excel_row}")

        validation = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTIONS) + '"', allow_blank=False)
        ws.add_data_validation(validation)
        validation.add(f"A2:A{len(rows) + 1}")
        ws.conditional_formatting.add(
            f"A2:S{len(rows) + 1}",
            FormulaRule(formula=['$A2="Resolved"'], fill=PatternFill("solid", fgColor="E2F0D9")),
        )

        # Summary/dashboard sheet.
        summary.sheet_view.showGridLines = False
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 18
        summary.column_dimensions["C"].width = 92
        summary["A1"] = "CCG Manual-Input Review Checklist"
        summary["A1"].font = Font(name="Aptos Display", size=20, bold=True, color="17365D")
        summary.merge_cells("A1:C1")
        summary["A3"] = "Purpose"
        summary["A3"].font = Font(bold=True, color="FFFFFF")
        summary["A3"].fill = PatternFill("solid", fgColor="17365D")
        summary["B3"] = "Value"
        summary["B3"].font = Font(bold=True, color="FFFFFF")
        summary["B3"].fill = PatternFill("solid", fgColor="17365D")
        summary["C3"] = "Meaning"
        summary["C3"].font = Font(bold=True, color="FFFFFF")
        summary["C3"].fill = PatternFill("solid", fgColor="17365D")
        summary_rows = [
            ("Cards requiring manual input", len(rows), "Every row is hash-pinned to the reviewed website text and Lua script."),
            ("Manual-ruling cases", sum(item["static_verdict"] == "MANUAL_RULING" for item in rows), "Needs an authoritative interpretation and/or interactive Omega evidence."),
            ("Engine limitations", sum(item["static_verdict"] == "UNSUPPORTED" for item in rows), "The printed behavior is not faithfully expressible with the bundled Omega API."),
            ("Resolved", f'=COUNTIF(\'{checklist_title}\'!A:A,"Resolved")', "Updates automatically as checklist statuses change."),
            ("Remaining", f'=COUNTA(\'{checklist_title}\'!D:D)-1-COUNTIF(\'{checklist_title}\'!A:A,"Resolved")', "Updates automatically."),
        ]
        for row, values in enumerate(summary_rows, 4):
            for col, value in enumerate(values, 1):
                cell = summary.cell(row, col, value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            summary.cell(row, 1).font = Font(bold=True)
        summary["A11"] = "How to use"
        summary["A11"].font = Font(name="Aptos Display", size=14, bold=True, color="17365D")
        instructions = [
            "1. Filter the checklist by Workstream, Static verdict, or Status.",
            "2. Read the full printed effect. Red bold text marks the clause selected from the audit evidence.",
            "3. Record the adopted ruling or engine decision in Reviewer notes / adopted ruling.",
            "4. Change the website card text or Lua only after that decision, then rerun the full CCG QA suite.",
            "5. Mark Resolved only after the evidence/acceptance condition has been met.",
        ]
        summary["A12"] = "\n".join(instructions)
        summary.merge_cells("A12:C16")
        summary["A12"].alignment = Alignment(vertical="top", wrap_text=True)
        summary["A12"].fill = PatternFill("solid", fgColor="D9EAF7")
        summary["A12"].border = border
        summary["A18"] = "Legend"
        summary["A18"].font = Font(name="Aptos Display", size=14, bold=True, color="17365D")
        legend = [
            ("Red bold text", "Problematic or ruling-sensitive clause inside the full printed effect."),
            ("Yellow cell", "Exact extracted issue-effect text."),
            ("Orange cell", "Explanation of what is wrong or unknown."),
            ("Purple verdict", "Known bundled-Omega engine limitation."),
            ("Green row", "Status has been changed to Resolved."),
        ]
        for row, (label, meaning) in enumerate(legend, 19):
            summary.cell(row, 1, label).font = Font(bold=True)
            summary.cell(row, 2, meaning)
            summary.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            for col in range(1, 4):
                summary.cell(row, col).border = border
                summary.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

        # Save while thumbnail files still exist; openpyxl reads them during save.
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(workbook_path)

    apply_excel_native_highlights(workbook_path, checklist_title, highlight_spans_by_row)
    return wb, {
        "cards": len(rows),
        "manual_rulings": sum(item["static_verdict"] == "MANUAL_RULING" for item in rows),
        "engine_limitations": sum(item["static_verdict"] == "UNSUPPORTED" for item in rows),
    }


def validate(path: Path) -> dict[str, Any]:
    skips = json.loads(SKIPS_PATH.read_text(encoding="utf-8"))
    expected_cards = len(skips["unresolved"])
    checklist_title = f"{expected_cards} Card Checklist"
    wb = load_workbook(path, read_only=False, rich_text=True, data_only=False)
    errors: list[str] = []
    expected_sheets = ["Summary", checklist_title]
    if wb.sheetnames != expected_sheets:
        errors.append(f"sheet names differ: {wb.sheetnames}")
    ws = wb[checklist_title]
    card_rows = ws.max_row - 1
    if card_rows != expected_cards:
        errors.append(f"checklist has {card_rows} cards, expected {expected_cards}")
    ids = [ws.cell(row, 5).value for row in range(2, ws.max_row + 1)]
    if len(ids) != len(set(ids)):
        errors.append("duplicate passcodes in checklist")
    if any(not ws.cell(row, 7).value for row in range(2, ws.max_row + 1)):
        errors.append("one or more rows has no full effect text")
    if any(not ws.cell(row, 8).value for row in range(2, ws.max_row + 1)):
        errors.append("one or more rows has no highlighted issue text")
    if any(not ws.cell(row, 11).value for row in range(2, ws.max_row + 1)):
        errors.append("one or more rows has no resolution action")
    rich_rows = 0
    full_cell_highlights = 0
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, 7).value
        if isinstance(value, CellRichText) and any(isinstance(block, TextBlock) for block in value):
            rich_rows += 1
        elif ws.cell(row, 7).font.bold and ws.cell(row, 7).font.color is not None:
            # Excel stores formatting on the cell itself when the selected issue
            # span is the complete printed text.
            full_cell_highlights += 1
    highlighted_rows = rich_rows + full_cell_highlights
    if highlighted_rows != expected_cards:
        errors.append(f"workbook has {highlighted_rows} highlighted effect rows, expected {expected_cards}")
    image_count = len(getattr(ws, "_images", []))
    if image_count != expected_cards:
        errors.append(f"workbook has {image_count} embedded card images, expected {expected_cards}")
    return {
        "path": str(path),
        "cards": card_rows,
        "unique_passcodes": len(set(ids)),
        "embedded_images": image_count,
        "rich_text_effect_rows": rich_rows,
        "full_cell_effect_highlights": full_cell_highlights,
        "highlighted_effect_rows": highlighted_rows,
        "sheets": wb.sheetnames,
        "errors": errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()
    skips = json.loads(SKIPS_PATH.read_text(encoding="utf-8"))
    workbook_path = output_path(len(skips["unresolved"]))
    counts: dict[str, int] = {}
    if not args.validate_only:
        _, counts = build_workbook()
    result = validate(workbook_path)
    print(json.dumps({**counts, **result}, indent=2))
    return 1 if result["errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
