#!/usr/bin/env python3
"""Build an image-backed Excel checklist from the fresh effect-by-effect audit."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import tempfile
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.text import InlineFont
from PIL import Image as PILImage


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_AUDIT = ROOT / "scripts/output/ccg_effect_by_effect_official_reference_audit.json"
DEFAULT_CARDS = ROOT / "src/data/cards.json"
DEFAULT_OUTPUT = ROOT / "scripts/output/CCG_remaining_rulings_engine_limitations_checklist.xlsx"

STOP_WORDS = {
    "a", "an", "and", "as", "at", "be", "by", "can", "card", "cards", "do",
    "effect", "effects", "for", "from", "if", "in", "is", "it", "of", "on",
    "or", "that", "the", "then", "this", "to", "until", "when", "with", "you",
    "your",
}


def tokens(value: str) -> set[str]:
    return {
        word for word in re.findall(r"[a-z0-9]+", value.lower())
        if len(word) > 1 and word not in STOP_WORDS
    }


def split_clauses(text: str) -> list[tuple[int, int, str]]:
    spans: list[tuple[int, int, str]] = []
    for match in re.finditer(r"[^\n]+?(?:[.!?](?=\s|$)|(?=\n)|$)", text, re.S):
        clause = match.group(0)
        if clause.strip():
            spans.append((match.start(), match.end(), clause))
    return spans


def best_issue_span(full_text: str, printed_clause: str) -> tuple[int, int] | None:
    wanted = tokens(printed_clause)
    if not wanted:
        return None
    best: tuple[float, int, int] | None = None
    for start, end, clause in split_clauses(full_text):
        found = tokens(clause)
        if not found:
            continue
        score = len(wanted & found) / max(1, len(wanted | found))
        candidate = (score, start, end)
        if best is None or candidate > best:
            best = candidate
    if best is None or best[0] < 0.18:
        return None
    return best[1], best[2]


def rich_card_text(full_text: str, printed_clause: str):
    span = best_issue_span(full_text, printed_clause)
    normal = InlineFont(rFont="Aptos", sz=10)
    issue = InlineFont(rFont="Aptos", sz=10, b=True, color="9C0006")
    rich = CellRichText()
    if not span:
        rich.append(TextBlock(normal, full_text))
        if full_text:
            rich.append(TextBlock(normal, "\n\n"))
        rich.append(TextBlock(issue, f"[PROBLEMATIC PRINTED CLAUSE IS MISSING: {printed_clause}]"))
        return rich
    start, end = span
    if start:
        rich.append(TextBlock(normal, full_text[:start]))
    rich.append(TextBlock(issue, full_text[start:end]))
    if end < len(full_text):
        rich.append(TextBlock(normal, full_text[end:]))
    return rich


def classify(issue: str) -> str:
    value = issue.lower()
    if any(term in value for term in (
        "ambiguous", "ambiguity", "card text", "wording", "ruling", "does not say",
        "no printed", "printed text omits", "needs a ruling", "text needs",
    )):
        return "Ruling / text ambiguity"
    if any(term in value for term in (
        "engine", "no generic", "no universal", "cannot generically", "helper",
        "unsupported", "no exact official", "no official", "cannot exactly",
        "not fully implemented", "no safe card-local", "requires engine",
    )):
        return "Engine limitation"
    return "Implementation limitation"


def resolution(issue: str, issue_type: str) -> str:
    if issue_type == "Ruling / text ambiguity":
        return (
            "Publish an authoritative ruling or erratum that resolves the ambiguity described at left. "
            "Update the website text and Lua to that ruling, add focused legality/resolution tests, and "
            "rerun the effect-by-effect audit before marking this row Resolved."
        )
    if issue_type == "Engine limitation":
        return (
            "Implement and expose the missing Omega/EDOPro engine or shared-helper capability described "
            "at left, then replace the current approximation and add focused regression tests. If engine "
            "support will not be added, rewrite the printed effect to an officially supported operation. "
            "Re-audit the effect before marking this row Resolved."
        )
    return (
        "Specify the exact intended edge-case behavior described at left, replace the custom approximation "
        "with a certifiable implementation, add focused gameplay tests, and rerun the effect-by-effect "
        "audit before marking this row Resolved."
    )


def resolve_image(card: dict) -> Path | None:
    relative = str(card.get("image") or "").lstrip("/")
    candidates = [ROOT / "public" / relative, ROOT / relative]
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved.is_file() and (resolved == ROOT or ROOT in resolved.parents):
            return resolved
    return None


def make_thumbnail(source: Path, destination: Path) -> None:
    with PILImage.open(source) as image:
        image = image.convert("RGB")
        image.thumbnail((118, 168), PILImage.Resampling.LANCZOS)
        canvas = PILImage.new("RGB", (118, 168), "white")
        canvas.paste(image, ((118 - image.width) // 2, (168 - image.height) // 2))
        canvas.save(destination, "JPEG", quality=82, optimize=True)


def reference_text(effect: dict) -> str:
    refs = []
    for ref in effect.get("official_references", []):
        label = f"{ref.get('card_name') or 'Unknown'} ({ref.get('card_id')}) — {ref.get('script')}"
        structure = ref.get("matched_structure")
        refs.append(f"{label}\n{structure}" if structure else label)
    return "\n\n".join(refs)


def load_rows(audit_path: Path, cards_path: Path) -> list[dict]:
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    cards = json.loads(cards_path.read_text(encoding="utf-8"))
    cards_by_passcode: dict[int, dict] = {}
    for card in cards:
        passcode = int(card["passcode"])
        if passcode in cards_by_passcode:
            raise ValueError(f"Duplicate card passcode in {cards_path}: {passcode}")
        cards_by_passcode[passcode] = card
    rows = []
    for audited in audit.get("cards", []):
        card = cards_by_passcode.get(int(audited["passcode"]), {})
        for effect in audited.get("effects", []):
            if effect.get("verdict") != "CUSTOM_GAP":
                continue
            if not card:
                raise ValueError(
                    f"Audit CUSTOM_GAP card {audited.get('passcode')} is absent from {cards_path}"
                )
            issue = effect.get("issue") or "The current audit identifies an unresolved ruling or engine limitation."
            printed_clause = effect.get("printed_clause") or ""
            if not printed_clause.strip():
                raise ValueError(
                    f"CUSTOM_GAP {audited['ordinal']} effect {effect.get('effect_index')} has no printed_clause"
                )
            if not str(card.get("text") or "").strip():
                raise ValueError(f"CUSTOM_GAP {audited['ordinal']} has no website card text")
            issue_type = classify(issue)
            rows.append({
                "ordinal": audited["ordinal"],
                "passcode": audited["passcode"],
                "name": audited["name"],
                "archetype": card.get("archetype") or "",
                "category": card.get("category") or "",
                "full_text": card.get("text") or "",
                "printed_clause": printed_clause,
                "issue": issue,
                "issue_type": issue_type,
                "resolution": resolution(issue, issue_type),
                "lua": effect.get("lua_implementation") or "",
                "references": reference_text(effect),
                "script": f"public/CCG Downloads/CCG_Scripts/c{audited['passcode']}.lua",
                "source": audited.get("source_batch") or "",
                "image": resolve_image(card),
            })
    return rows


def display_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT))
    except ValueError:
        return str(path.resolve())


def build_workbook(rows: list[dict], output: Path, audit_path: Path, cards_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    issues = workbook.create_sheet("Checklist")

    navy = "17365D"
    blue = "D9EAF7"
    yellow = "FFF2CC"
    red = "F4CCCC"
    green = "D9EAD3"
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    unique_cards = len({row["passcode"] for row in rows})
    type_counts = Counter(row["issue_type"] for row in rows)
    summary["A1"] = "CCG Remaining Rulings / Engine Limitations"
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary.merge_cells("A1:D1")
    summary.append(["Current unresolved effect records", len(rows), "Unique cards", unique_cards])
    summary.append(["Audit source", display_path(audit_path), "Card source", display_path(cards_path)])
    summary.append([])
    summary.append(["Issue type", "Count", "Checklist meaning", ""])
    for issue_type, count in sorted(type_counts.items()):
        summary.append([issue_type, count, "Open until the listed resolution is implemented and re-audited.", ""])
    summary.append([])
    summary.append(["Status choices", "☐ Open / Ruling needed / Engine work / Text change / Resolved / Won't fix", "", ""])
    summary.append(["Highlighting", "The problematic printed effect is isolated in a yellow cell. Its matching sentence is bold red in the full website text; if the required clause is absent, a bold red missing-clause marker is appended.", "", ""])
    summary.append(["Scope note", "This workbook is a point-in-time rendering of CUSTOM_GAP records in the selected audit. Regenerate it whenever the audit changes.", "", ""])
    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for cell in summary[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 55
    summary.column_dimensions["C"].width = 50
    summary.column_dimensions["D"].width = 18
    summary.freeze_panes = "A2"

    headers = [
        "Status", "Ordinal", "Passcode", "Card Image", "Card Name", "Archetype", "Category",
        "Full Website Card Effect Text (issue bold red)", "Problematic Effect Text", "Issue Type",
        "What Is Wrong", "What Needs To Be Done", "Current Lua / Approximation",
        "Official Omega Reference(s)", "Canonical Script", "Audit Source", "Reviewer Notes",
    ]
    issues.append(headers)
    for cell in issues[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    issues.row_dimensions[1].height = 38
    issues.freeze_panes = "A2"
    issues.auto_filter.ref = f"A1:Q{max(2, len(rows) + 1)}"

    status_validation = DataValidation(
        type="list",
        formula1='"☐ Open,Ruling needed,Engine work,Text change,Resolved,Won\'t fix"',
        allow_blank=False,
    )
    issues.add_data_validation(status_validation)

    temp_dir = Path(tempfile.mkdtemp(prefix="ccg_xlsx_thumbs_"))
    try:
        for index, item in enumerate(rows, start=2):
            values = [
                "☐ Open", item["ordinal"], item["passcode"], "", item["name"], item["archetype"],
                item["category"], "", item["printed_clause"], item["issue_type"], item["issue"],
                item["resolution"], item["lua"], item["references"], item["script"], item["source"], "",
            ]
            for column, value in enumerate(values, start=1):
                cell = issues.cell(index, column, value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
            issues.cell(index, 8).value = rich_card_text(item["full_text"], item["printed_clause"])
            issues.cell(index, 9).fill = PatternFill("solid", fgColor=yellow)
            issues.cell(index, 9).font = Font(bold=True, color="9C0006")
            issues.cell(index, 10).fill = PatternFill("solid", fgColor=blue)
            issues.row_dimensions[index].height = 132

            image_path = item["image"]
            if image_path:
                thumbnail = temp_dir / f"{item['passcode']}_{index}.jpg"
                make_thumbnail(image_path, thumbnail)
                picture = XLImage(thumbnail)
                picture.width = 88
                picture.height = 126
                picture.anchor = f"D{index}"
                issues.add_image(picture)
                issues.cell(index, 5).hyperlink = image_path.as_uri()
                issues.cell(index, 5).font = Font(color="0563C1", underline="single")
            else:
                issues.cell(index, 4).value = "Image missing"
                issues.cell(index, 4).fill = PatternFill("solid", fgColor=red)

        if rows:
            status_validation.add(f"A2:A{len(rows) + 1}")
            issues.conditional_formatting.add(
                f"A2:A{len(rows) + 1}",
                FormulaRule(formula=['LEFT(A2,8)="Resolved"'], fill=PatternFill("solid", fgColor=green)),
            )
        widths = {
            "A": 16, "B": 9, "C": 13, "D": 16, "E": 30, "F": 20, "G": 12,
            "H": 58, "I": 48, "J": 24, "K": 58, "L": 58, "M": 58, "N": 55,
            "O": 46, "P": 27, "Q": 35,
        }
        for column, width in widths.items():
            issues.column_dimensions[column].width = width

        for index, item in enumerate(rows, start=2):
            script_path = (ROOT / item["script"]).resolve()
            if script_path.is_file():
                issues.cell(index, 15).hyperlink = script_path.as_uri()
                issues.cell(index, 15).font = Font(color="0563C1", underline="single")
            source_path = (ROOT / "scripts/output/effect_audit_agents" / item["source"]).resolve()
            if source_path.is_file():
                issues.cell(index, 16).hyperlink = source_path.as_uri()
                issues.cell(index, 16).font = Font(color="0563C1", underline="single")

        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument("--cards", type=Path, default=DEFAULT_CARDS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    rows = load_rows(args.audit, args.cards)
    missing_images = [row for row in rows if not row["image"]]
    if missing_images:
        details = ", ".join(f"{row['ordinal']} {row['name']}" for row in missing_images)
        raise FileNotFoundError(f"Website images missing for CUSTOM_GAP rows: {details}")
    build_workbook(rows, args.output, args.audit, args.cards)
    print(json.dumps({
        "output": str(args.output.resolve()),
        "unresolved_effects": len(rows),
        "unique_cards": len({row['passcode'] for row in rows}),
        "missing_images": len(missing_images),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
