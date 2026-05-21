from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


REPO_ROOT = Path(__file__).resolve().parents[1]
CARDS_PATH = REPO_ROOT / "src" / "data" / "cards.json"
OUTPUT_DIR = REPO_ROOT / "scripts" / "output"
OUTPUT_PATH = OUTPUT_DIR / "YGO_Omega_Card_Checklist.xlsx"

CARDS_SHEET = "Cards"
SUMMARY_SHEET = "Summary"
STATUS_FIELDS = ("Coded", "Tested", "Confirmed")
DEFAULT_STATUS = "No"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_SIDE = Side(style="thin", color="D9E2F3")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
YES_FILL = PatternFill("solid", fgColor="C6E0B4")
NO_FILL = PatternFill("solid", fgColor="F4CCCC")
TITLE_FONT = Font(size=14, bold=True)
SUBTITLE_FONT = Font(size=11, italic=True, color="666666")


def load_cards() -> list[dict[str, Any]]:
    return json.loads(CARDS_PATH.read_text(encoding="utf-8"))


def split_set(set_value: str | None) -> tuple[str, str, str]:
    if not set_value:
        return "", "", ""

    match = re.match(r"^([A-Z0-9]+)-(\d+)\s+(.*)$", set_value)
    if not match:
        return "", "", set_value

    return match.group(1), match.group(2), match.group(3)


def type_line(card: dict[str, Any]) -> str:
    if card.get("category") == "Monster":
        species = " / ".join(card.get("monsterType") or [])
        subtypes = " / ".join(card.get("cardTypes") or [])
        return " / ".join(part for part in (species, subtypes) if part)

    icon = card.get("icon")
    category = card.get("category")
    if icon and category:
        return f"{category} / {icon}"
    return str(category or "")


def load_existing_progress() -> dict[str, dict[str, str]]:
    if not OUTPUT_PATH.exists():
        return {}

    workbook = load_workbook(OUTPUT_PATH)
    if CARDS_SHEET not in workbook.sheetnames:
        return {}

    sheet = workbook[CARDS_SHEET]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1) if cell.value}
    required = {"ID", *STATUS_FIELDS, "Notes"}
    if not required.issubset(headers):
        return {}

    progress: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        card_id = row[headers["ID"] - 1]
        if not card_id:
            continue

        entry = {field: str(row[headers[field] - 1] or DEFAULT_STATUS) for field in STATUS_FIELDS}
        entry["Notes"] = str(row[headers["Notes"] - 1] or "")
        progress[str(card_id)] = entry

    return progress


def build_rows(cards: list[dict[str, Any]], existing_progress: dict[str, dict[str, str]]) -> list[list[Any]]:
    rows: list[list[Any]] = []

    for card in cards:
        card_id = str(card.get("id") or "")
        set_code, set_number, set_name = split_set(card.get("set"))
        progress = existing_progress.get(card_id, {})

        row = [
            card_id,
            str(card.get("name") or ""),
            set_code,
            set_number,
            set_name,
            str(card.get("category") or ""),
            type_line(card),
            str(card.get("archetype") or ""),
            str((card.get("timestamps") or {}).get("added") or ""),
            str(card.get("image") or ""),
            progress.get("Coded", DEFAULT_STATUS),
            progress.get("Tested", DEFAULT_STATUS),
            progress.get("Confirmed", DEFAULT_STATUS),
            "",
            progress.get("Notes", ""),
        ]
        rows.append(row)

    return rows


def style_header(sheet, header_row: int, width_map: dict[str, int]) -> None:
    for cell in sheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value in width_map:
            sheet.column_dimensions[cell.column_letter].width = width_map[cell.value]


def apply_card_sheet_formatting(sheet, row_count: int) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:O{row_count}"

    for row in sheet.iter_rows(min_row=2, max_row=row_count, min_col=1, max_col=15):
        for cell in row:
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_index in range(2, row_count + 1):
        sheet[f"N{row_index}"] = f'=IF(AND($K{row_index}="Yes",$L{row_index}="Yes",$M{row_index}="Yes"),"Yes","No")'

    status_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    sheet.add_data_validation(status_validation)
    status_validation.add(f"K2:M{row_count}")

    green_rule = FormulaRule(formula=['K2="Yes"'], fill=YES_FILL)
    red_rule = FormulaRule(formula=['K2="No"'], fill=NO_FILL)
    ready_green_rule = FormulaRule(formula=['N2="Yes"'], fill=YES_FILL)
    ready_red_rule = FormulaRule(formula=['N2="No"'], fill=NO_FILL)

    for column in ("K", "L", "M"):
        sheet.conditional_formatting.add(f"{column}2:{column}{row_count}", green_rule)
        sheet.conditional_formatting.add(f"{column}2:{column}{row_count}", red_rule)

    sheet.conditional_formatting.add(f"N2:N{row_count}", ready_green_rule)
    sheet.conditional_formatting.add(f"N2:N{row_count}", ready_red_rule)

    table = Table(displayName="OmegaChecklist", ref=f"A1:O{row_count}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def write_summary_sheet(workbook: Workbook, cards: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(SUMMARY_SHEET)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    unique_sets: list[tuple[str, str]] = []
    seen_sets: set[tuple[str, str]] = set()

    for card in cards:
        set_code, _set_number, set_name = split_set(card.get("set"))
        key = (set_code, set_name)
        if key not in seen_sets:
            seen_sets.add(key)
            unique_sets.append(key)

    sheet["A1"] = "YGO Omega Card Checklist"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"Generated from src/data/cards.json on {generated_at}"
    sheet["A2"].font = SUBTITLE_FONT

    summary_rows = [
        ("Total Cards", '=COUNTA(Cards!$A:$A)-1'),
        ("Coded", '=COUNTIF(Cards!$K:$K,"Yes")'),
        ("Tested", '=COUNTIF(Cards!$L:$L,"Yes")'),
        ("Confirmed", '=COUNTIF(Cards!$M:$M,"Yes")'),
        ("Ready For Omega", '=COUNTIF(Cards!$N:$N,"Yes")'),
    ]

    start_row = 4
    for offset, (label, formula) in enumerate(summary_rows):
        row = start_row + offset
        sheet[f"A{row}"] = label
        sheet[f"B{row}"] = formula
        sheet[f"C{row}"] = f'=IF($B$4=0,0,B{row}/$B$4)'

    for cell in sheet[f"A{start_row}:C{start_row + len(summary_rows) - 1}"]:
        for item in cell:
            item.border = CELL_BORDER

    for row in range(start_row, start_row + len(summary_rows)):
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"C{row}"].number_format = "0.0%"

    table_start = 11
    headers = ["Set Code", "Set Name", "Total", "Coded", "Tested", "Confirmed", "Ready"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=table_start, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center")

    for offset, (set_code, set_name) in enumerate(unique_sets, start=1):
        row = table_start + offset
        sheet[f"A{row}"] = set_code
        sheet[f"B{row}"] = set_name
        sheet[f"C{row}"] = f'=COUNTIF(Cards!$C:$C,A{row})'
        sheet[f"D{row}"] = f'=COUNTIFS(Cards!$C:$C,A{row},Cards!$K:$K,"Yes")'
        sheet[f"E{row}"] = f'=COUNTIFS(Cards!$C:$C,A{row},Cards!$L:$L,"Yes")'
        sheet[f"F{row}"] = f'=COUNTIFS(Cards!$C:$C,A{row},Cards!$M:$M,"Yes")'
        sheet[f"G{row}"] = f'=COUNTIFS(Cards!$C:$C,A{row},Cards!$N:$N,"Yes")'
        for column in range(1, 8):
            sheet.cell(row=row, column=column).border = CELL_BORDER

    sheet.freeze_panes = "A11"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 28
    for column in ("C", "D", "E", "F", "G"):
        sheet.column_dimensions[column].width = 14


def build_workbook(cards: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    existing_progress = load_existing_progress()
    rows = build_rows(cards, existing_progress)

    sheet = workbook.create_sheet(CARDS_SHEET)
    headers = [
        "ID",
        "Name",
        "Set Code",
        "Set Number",
        "Set Name",
        "Category",
        "Type Line",
        "Archetype",
        "Added",
        "Image",
        "Coded",
        "Tested",
        "Confirmed",
        "Ready For Omega",
        "Notes",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    style_header(
        sheet,
        1,
        {
            "ID": 14,
            "Name": 38,
            "Set Code": 12,
            "Set Number": 12,
            "Set Name": 26,
            "Category": 12,
            "Type Line": 34,
            "Archetype": 22,
            "Added": 14,
            "Image": 44,
            "Coded": 12,
            "Tested": 12,
            "Confirmed": 12,
            "Ready For Omega": 16,
            "Notes": 36,
        },
    )
    apply_card_sheet_formatting(sheet, len(rows) + 1)
    write_summary_sheet(workbook, cards)

    return workbook


def main() -> int:
    cards = load_cards()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(cards)
    workbook.save(OUTPUT_PATH)
    print(f"Wrote checklist workbook: {OUTPUT_PATH}")
    print(f"Cards included: {len(cards)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
